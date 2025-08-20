# -*- coding: utf-8 -*-
"""
Created on Fri May 10 11:43:50 2024

@author: celinep
"""

from openpyxl import load_workbook, Workbook


class ExcelTIMES:
    """ Master class for dealing with Excel files """

    def __init__(self, wb_name:str, ws_name:str, data_only:bool, delete_old=True):
        """
        parameters:
        data_only: may erase existing formulas in the workbook
        delete_old: when loading the worksheet, erase existing data
        """

        self.wb_name = wb_name
        try:
            self.wb = load_workbook(wb_name , data_only=data_only)
        except FileNotFoundError:
            self.wb = Workbook()
        try:
            self.ws = self.wb[ws_name]
        except KeyError:
            self.wb.create_sheet(ws_name)
            self.ws = self.wb[ws_name]
        if delete_old:
            self.ws.delete_rows(1, self.ws.max_row)
        self.row_nb = 1

    def close(self):
        """ save and close the workbook """
        self.wb.save(self.wb_name)
        self.wb.close()

    def write_header(self, name, header:list):
        """ Write the header of tables for TIMES use """
        self.ws.cell(self.row_nb, 1, value=name)
        for i, value in enumerate(header):
            for j, value2 in enumerate(value):
                self.ws.cell(column=j+1, row=self.row_nb+1+i, value=value2)
        self.row_nb += 1 + len(header)


    def sow(self, stages:dict, sows:dict):
        """ Write the table for TIMES to define a SP
        Parameters:
            - stages: a dict associating each stage number with the year it starts
            - sows: a dict associating each scenario number to its probability
        """
        self.write_header("~TFM_INS", [["Attribute", "STAGE", "SOW", "Value"]])
        # Define the decision period
        for stage, year in stages.items():
            self.ws.cell(column=1, row=self.row_nb, value="SW_START")
            self.ws.cell(column=2, row=self.row_nb, value=stage)
            self.ws.cell(column=4, row=self.row_nb, value=year)
            self.row_nb += 1
        # Define the nb of SOW per stage (only 2 stages considered here, to be changed otherwise)
        self.ws.cell(column=1, row=self.row_nb, value="SW_SUBS")
        self.ws.cell(column=2, row=self.row_nb, value=1)
        self.ws.cell(column=3, row=self.row_nb, value=1)
        self.ws.cell(column=4, row=self.row_nb, value=len(sows.keys()))
        self.row_nb += 1
        # Define the SOW and their probabilities
        for sow, p in sows.items():
            self.ws.cell(column=1, row=self.row_nb, value="SW_SPROB")
            self.ws.cell(column=2, row=self.row_nb, value=2)
            self.ws.cell(column=3, row=self.row_nb, value=sow)
            self.ws.cell(column=4, row=self.row_nb, value=p)
            self.row_nb += 1
        self.row_nb += 2


    def write_scenarios(self, scenarios, new_scenarios=True):
        """To be updated """
        col_level = 1
        col_var = 2
        col_full_scenarios = 3
        for row in range(1, 50):
            if self.ws.cell(row=row, column=col_level).value is None:
                var = self.ws.cell(row=row, column=col_var).value
            else:
                level = self.ws.cell(row=row, column=col_level).value
                txt = "0,99,100"
                for j, s in enumerate(scenarios):
                    if (f"{var}_{level.lower()}" in str(s) and not new_scenarios)\
                        or (str(s) in str(self.ws.cell(row=row,
                                                        column=col_full_scenarios).value).split(",") and new_scenarios):
                        txt += f",{j+1}"
                self.ws.cell(column=col_var, row=row, value=txt)
                if not new_scenarios:
                    self.ws.cell(column=col_full_scenarios, row=row, value=txt)

    def write_scenarios_par_2s(self, pairs):
        """ To be updated """
        # scenarios
        col_L, line_L = 3,5
        col_H, line_H = 3, 40
        col_level = 13
        col_var = 14
        col_full_scenarios = 15

        for row in range(1, 50):
            if self.ws.cell(row=row, column=col_level).value is None:
                var = self.ws.cell(row=row, column=col_var).value
            else:
                level = self.ws.cell(row=row, column=col_level).value

                for k, (i,j) in pairs.items():
                    txt = "0,99,100"
                    txt = ""
                    first = True

                    # ex: 1, (21,44)
                    print(k,i,j)
                    for n,s in enumerate((i,j)):
                        if str(s) in str(self.ws.cell(row=row,
                                                      column=col_full_scenarios).value).split(","):
                            if not first:
                                txt += ","
                            txt += f"{n+1}"
                            first = False

                    l_base = line_H  if level == "HIGH" else line_L

                    for c in range(col_L+1, col_L+8):
                        if self.ws.cell(row=l_base, column=c).value == var:
                            self.ws.cell(column=c, row=l_base+k, value=txt)
                            break


    def write_scenarios_par(self, K_clusters:dict, N:int):
        """ Write the parametric sheet to associate each case with a number of scenarios for the reduced SP.
        Parameters:
        - K_clusters: clustering of the scenarios like {1:[1,2,5], 3:[3,6]} with K values keys
        - N: the total number of scenarios of the full SP
        """

        target_proba = "PROBA"
        target_level = ["HIGH", "LOW", "MID"]
        cells_target = {t:None for t in target_level+[target_proba]}

        dict_case_to_K = {}
        ws_uncertainties = self.wb["Uncertainties"]
        row_case = 6
        print(f"reading {ws_uncertainties.cell(row=row_case, column=1).value}...")
        if ws_uncertainties.cell(row=row_case, column=1).value != "Case":
            raise ValueError
        for row in range(row_case+1, 100):
            if ws_uncertainties.cell(row=row, column=1).value is None:
                break
            dict_case_to_K[int(ws_uncertainties.cell(row=row, column=2).value)] \
                = int(ws_uncertainties.cell(row=row, column=1).value)
        nb_cases = len(dict_case_to_K)

        for row in self.ws.iter_rows(values_only=False):
            for cell in row:
                if cell.value in target_level+[target_proba]:
                    row_idx = cell.row
                    col_idx = cell.column
                    left_ok = (col_idx > 1 and self.ws.cell(row=row_idx+1,
                                                            column=col_idx-1).value == "scenarios")
                    top_ok  = (row_idx > 1 and self.ws.cell(row=row_idx-1,
                                                             column=col_idx+1).value in ["cases",
                                                                                          "uncertainties"])
                    if left_ok and top_ok:
                        cells_target[cell.value] = (cell.row, cell.column)
                    if None not in cells_target.values():
                        break
            if None not in cells_target.values():
                break

        for K, cluster in K_clusters.items():
            # Write probability of each scenario of the approximate problem for all cases
            s = 1
            for list_scenarios in cluster.values():
                proba_cluster = len(list_scenarios) / N
                self.ws.cell(column=cells_target[target_proba][1]+dict_case_to_K[K],
                            row=cells_target[target_proba][0]+s,
                            value=proba_cluster)
                s += 1

            # Write the scenarios associated with the right level of uncertainties
            col_level, col_uncert, col_fullSP = 1, 2, 3
            representatives = cluster.keys()
            for row in range(1,100):
                if self.ws.cell(row=row, column=col_level).value is None:
                    if self.ws.cell(row=row+1, column=col_level).value is None:
                        break
                    uncertainty = self.ws.cell(row=row, column=col_uncert).value
                    # We locate the uncertainty
                else:
                    level = self.ws.cell(row=row, column=col_level).value
                    # We locate the level of uncertainty
                    txt = ""
                    first = True
                    # For each scenario, we redefine the numerotation of scenarios based on new K
                    for j, repre in enumerate(representatives):
                        if str(repre) in str(self.ws.cell(row=row, column=col_fullSP).value).split(","):
                            if not first:
                                txt += ","
                            txt += f"{j+1}"
                            first = False
                    # For each case (which represent different K values)
                    for col_case in range(cells_target[level][1]+1, cells_target[level][1]+nb_cases):
                        if self.ws.cell(row=cells_target[level][0],
                                         column=col_case).value == uncertainty:
                            self.ws.cell(row=cells_target[level][0]+K,
                                         column=col_case, value=txt)
                            break
